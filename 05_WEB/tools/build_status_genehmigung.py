#!/usr/bin/env python3
"""Build parcel permission status from Excel permit tables and local WFS parcels."""

from __future__ import annotations

import json
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl


CADASTRE_GEOJSON = "../data/catastro_flurstueck.geojson"
OUTPUT_GEOJSON = "../data/status_genehmigung.geojson"
PERMIT_SHEETS = ("Eigentümer BL02", "Eigentümer BL03", "Eigentürmer BL02", "Eigentürmer BL03", "Liste")


def clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def normalize_text(value) -> str:
    return clean_text(value).casefold()


def normalize_header(value) -> str:
    text = unicodedata.normalize("NFKD", clean_text(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", text.casefold())


def normalize_flur(value) -> str:
    text = clean_text(value)
    match = re.search(r"\d+", text)
    return match.group(0) if match else text


def normalize_flurstueck(value) -> str:
    text = clean_text(value)
    if not text:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = text.replace(" ", "")
    text = re.sub(r"\.0$", "", text)
    return text


def parcel_key(gemarkung, flur, flurstueck) -> tuple[str, str, str]:
    return (normalize_text(gemarkung), normalize_flur(flur), normalize_flurstueck(flurstueck))


def is_truthy(value) -> bool:
    if value is True:
        return True
    if value is False or value is None:
        return False
    text = normalize_text(value)
    return text in {"true", "wahr", "ja", "yes", "x", "1", "✓"}


def format_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean_text(value)


def parse_date_value(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def header_value(header_index: dict[str, int], row, *names):
    for name in names:
        idx = header_index.get(normalize_header(name))
        if idx is not None and idx < len(row):
            return row[idx]
    return None


def values_for_header(headers: list[str], row, *names):
    normalized_names = {normalize_header(name) for name in names}
    for idx, header in enumerate(headers):
        if normalize_header(header) in normalized_names and idx < len(row):
            yield row[idx]


def find_header_row(ws):
    previous = None
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        headers = [clean_text(v) for v in row]
        normalized = {normalize_header(v) for v in headers if v}
        has_flurstueck = "flurstuck" in normalized or "flurstueck" in normalized
        if {"ampel", "flur"} <= normalized and has_flurstueck:
            merged = []
            for idx, header in enumerate(headers):
                if header:
                    merged.append(header)
                elif previous and idx < len(previous) and clean_text(previous[idx]) == "Kommentar":
                    merged.append("Kommentar")
                else:
                    merged.append("")
            return row_number, merged
        previous = headers
    return None, None


def ampel_status(value) -> str:
    text = normalize_text(value)
    if not text:
        return "nicht_informiert"
    if "keine kontaktdaten" in text:
        return "nicht_informiert"
    if "zustimmung" in text and "erteilt" in text:
        return "genehmigt"
    if "kontaktiert" in text:
        return "informiert_offen"
    return "informiert_offen"


def new_entry():
    return {
        "masts": set(),
        "owners": set(),
        "first_names": set(),
        "last_names": set(),
        "emails": set(),
        "phones": set(),
        "baulose": set(),
        "leitungen": set(),
        "work_types": set(),
        "remarks": set(),
        "info_dates": set(),
        "contact_dates": set(),
        "ampel_values": set(),
        "ampel_statuses": [],
        "permission_values": [],
        "rows": 0,
    }


def read_permit_rows(excel_path: Path) -> dict[tuple[str, str, str], dict]:
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    permits = defaultdict(new_entry)

    for sheet_name in PERMIT_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_row_number, headers = find_header_row(ws)
        if not headers:
            continue

        header_index = {}
        for idx, name in enumerate(headers):
            key = normalize_header(name)
            if key and key not in header_index:
                header_index[key] = idx

        blank_streak = 0
        for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_number <= header_row_number:
                continue
            if not any(v is not None for v in row[:12]):
                blank_streak += 1
                if blank_streak > 25:
                    break
                continue
            blank_streak = 0

            gemarkung = header_value(header_index, row, "Gemarkung")
            flur = header_value(header_index, row, "Flur")
            flurstueck = header_value(header_index, row, "Flurstueck", "Flurstück")
            key = parcel_key(gemarkung, flur, flurstueck)
            if not all(key):
                continue

            entry = permits[key]
            entry["rows"] += 1
            baulos = format_value(header_value(header_index, row, "Los", "Baulos"))
            entry["baulose"].add(baulos or sheet_name.replace("Eigentümer ", ""))
            entry["leitungen"].add(format_value(header_value(header_index, row, "Leitung")))
            entry["work_types"].add(format_value(header_value(header_index, row, "Art")))
            entry["masts"].add(format_value(header_value(header_index, row, "Mast")))

            first_name = format_value(header_value(header_index, row, "Vorname"))
            last_name = format_value(header_value(header_index, row, "Nachname"))
            if first_name:
                entry["first_names"].add(first_name)
            if last_name:
                entry["last_names"].add(last_name)

            owner = " ".join(part for part in [first_name, last_name] if part)
            if owner:
                entry["owners"].add(owner)

            for header in headers:
                header_norm = normalize_header(header)
                value = format_value(header_value(header_index, row, header))
                if not value:
                    continue
                if any(token in header_norm for token in ("mail", "email")):
                    entry["emails"].add(value)
                if any(token in header_norm for token in ("telefon", "tel", "phone", "mobil", "handy")):
                    entry["phones"].add(value)

            remark = format_value(header_value(header_index, row, "Bemerkung", "Kommentar"))
            if remark:
                entry["remarks"].add(remark)

            ampel = format_value(header_value(header_index, row, "Ampel"))
            entry["ampel_statuses"].append(ampel_status(ampel))
            if ampel:
                entry["ampel_values"].add(ampel)

            permission_value = header_value(header_index, row, "Zustimmung", "Erlaubnis erhalten")
            entry["permission_values"].append(is_truthy(permission_value))

            for header in headers:
                if normalize_header(header).startswith("fechainformacion"):
                    info_date = format_value(header_value(header_index, row, header))
                    if info_date:
                        entry["info_dates"].add(info_date)

            for contact_date in values_for_header(headers, row, "Datum"):
                parsed_date = parse_date_value(contact_date)
                if parsed_date:
                    entry["contact_dates"].add(parsed_date.isoformat())

    return permits


def status_for(entry: dict) -> str:
    if entry["ampel_statuses"]:
        if "genehmigt" in entry["ampel_statuses"]:
            return "genehmigt"
        if "informiert_offen" in entry["ampel_statuses"]:
            return "informiert_offen"
        return "nicht_informiert"
    if any(entry["permission_values"]):
        return "genehmigt"
    if not entry["info_dates"]:
        return "nicht_informiert"
    return "informiert_offen"


def merge_entry(target: dict, value: dict) -> None:
    for set_key in (
        "masts",
        "owners",
        "first_names",
        "last_names",
        "emails",
        "phones",
        "baulose",
        "leitungen",
        "work_types",
        "remarks",
        "info_dates",
        "contact_dates",
        "ampel_values",
    ):
        target[set_key].update(value[set_key])
    target["ampel_statuses"].extend(value["ampel_statuses"])
    target["permission_values"].extend(value["permission_values"])
    target["rows"] += value["rows"]


def sorted_masts(values):
    return sorted((v for v in values if v), key=lambda x: (len(x), x))


def german_date(value: str) -> str:
    parsed = parse_date_value(value)
    return parsed.strftime("%d.%m.%Y") if parsed else ""


def main() -> None:
    base = Path(__file__).resolve().parent
    excel_files = sorted((base / "../../04_PERMITS").resolve().glob("*.xlsx"))
    if not excel_files:
        raise FileNotFoundError("No .xlsx files found in 04_PERMITS")

    permits = {}
    for excel_path in excel_files:
        for key, value in read_permit_rows(excel_path).items():
            if key in permits:
                merge_entry(permits[key], value)
            else:
                permits[key] = value

    cadastre_path = (base / CADASTRE_GEOJSON).resolve()
    cadastre = json.loads(cadastre_path.read_text(encoding="utf-8"))
    features = []
    matched = 0

    for feature in cadastre["features"]:
        props = feature.get("properties", {})
        key = parcel_key(props.get("gemarkung"), props.get("flur"), props.get("flurstueck"))
        entry = permits.get(key)
        if not entry:
            continue
        matched += 1
        status = status_for(entry)
        latest_contact = max(entry["contact_dates"]) if entry["contact_dates"] else ""
        status_props = {
            **props,
            "status_genehmigung": status,
            "status_label": {
                "genehmigt": "Zustimmung erteilt",
                "nicht_informiert": "Nicht kontaktiert",
                "informiert_offen": "Kontaktiert",
            }[status],
            "ampel": "; ".join(sorted(entry["ampel_values"])),
            "baulos": ", ".join(sorted(v for v in entry["baulose"] if v)),
            "leitung": ", ".join(sorted(v for v in entry["leitungen"] if v)),
            "art": ", ".join(sorted(v for v in entry["work_types"] if v)),
            "masten": ", ".join(sorted_masts(entry["masts"])),
            "eigentuemer": "; ".join(sorted(entry["owners"])),
            "vorname": "; ".join(sorted(entry["first_names"])),
            "nachname": "; ".join(sorted(entry["last_names"])),
            "email": "; ".join(sorted(entry["emails"])),
            "telefon": "; ".join(sorted(entry["phones"])),
            "info_daten": ", ".join(sorted(entry["info_dates"])),
            "kontakt_daten": ", ".join(sorted(entry["contact_dates"])),
            "letzter_kontakt": latest_contact,
            "letzter_kontakt_label": german_date(latest_contact),
            "bemerkung": "; ".join(sorted(entry["remarks"])),
            "permit_rows": entry["rows"],
        }
        features.append({
            "type": "Feature",
            "properties": status_props,
            "geometry": feature["geometry"],
        })

    output_path = (base / OUTPUT_GEOJSON).resolve()
    output_path.write_text(
        json.dumps({
            "type": "FeatureCollection",
            "name": "status_genehmigung",
            "source": ", ".join(path.name for path in excel_files),
            "features": features,
        }, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )

    counts = defaultdict(int)
    for feature in features:
        counts[feature["properties"]["status_genehmigung"]] += 1
    print(f"Matched {matched} parcels. Wrote {len(features)} features to {output_path}")
    print(dict(counts))


if __name__ == "__main__":
    main()
